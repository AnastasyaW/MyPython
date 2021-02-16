import openpyxl

file = openpyxl.load_workbook("C:\\Users\\atcin\\Documents\\PythonDemo.xlxs")
sheet = file.active
# we will store the data we get from excell and then we 
#> will use in out test
dict = {}

cell = sheet.cell(row=1, comun=2)
print(cell.value)

# print back to excell file:
sheet.cell(row=2, column=2).value = "Rahul"
# the ma xnumber of the rows:
print(sheet.max_row)
# max colums nr
sheet.max_column
print(sheet["A5"].value)

# print the entire data
for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, colum=1).value == "testcase2":
        for x in range(2, sheet.max_column + 1):
            #here we add the value into the dictionary 
            dict[sheet.cell(row = 1, column = x).value] = sheet.cell(row=i, column=x).value)
